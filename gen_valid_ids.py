"""Generate Selmark_Valid_IDs.xlsx — respondents that passed QC (OK + REVIEW)."""
import json, os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import sys
sys.path.insert(0, str(Path(__file__).parent))
import quality_check as qc

src = Path(r'C:\Users\ajulve\Downloads\selmark_respuestas_full.json')
with open(src, encoding='utf-8') as f:
    data = json.load(f)
completed = [r for r in data if r.get('status') == 'complete']
results   = [qc.assess_response(r) for r in completed]

valid_rows = [
    (r['respondent_id'], res['verdict'], round(float(res['duration_min']), 1) if res['duration_min'] not in ('N/A', None) and str(res['duration_min']) != 'N/A' else 'N/A')
    for r, res in zip(completed, results)
    if res['verdict'] in ('OK', 'REVIEW')
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Valid Respondents'

header_fill = PatternFill('solid', fgColor='1F3864')
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
for col, text in enumerate(['Respondent ID', 'QC Status', 'Duration (min)'], start=1):
    cell = ws.cell(row=1, column=col, value=text)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

row_fills = [PatternFill('solid', fgColor='FFFFFF'), PatternFill('solid', fgColor='F2F2F2')]
for idx, (rid, verdict, dur) in enumerate(valid_rows, start=2):
    fill = row_fills[idx % 2]
    for col, val in enumerate([rid, verdict, dur], start=1):
        cell = ws.cell(row=idx, column=col, value=val)
        cell.font = Font(name='Arial', size=9)
        cell.fill = fill

ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 16

out = r'C:\Users\ajulve\Downloads\Selmark_Valid_IDs.xlsx'
wb.save(out)
print(f'Saved: {out}  ({len(valid_rows)} valid respondents)')
